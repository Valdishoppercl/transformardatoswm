import streamlit as st
import pandas as pd
import zipfile
import io
import re
from openpyxl import load_workbook

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Transformador de Indicadores WM", page_icon="📊", layout="wide")

# --- FUNCIONES DE APOYO ---
def parse_filters_text(text: str):
    if not isinstance(text, str): return {}
    t = text.replace("\n", " ").replace("\r", " ")
    def m_int(pattern):
        m = re.search(pattern, t, re.I)
        return int(m.group(1)) if m else None
    return {
        "Local ID": m_int(r"(?:local|sala|nodo)\s*(?:es|=|:)?\s*(\d{1,6})"),
        "Semana":   m_int(r"semana\s*(?:es|=|:)?\s*(\d{1,2})"),
        "Año":      m_int(r"año\s*(?:es|=|:)?\s*(\d{4})"),
    }

def to_number(v):
    if v is None or v == "" or pd.isna(v): return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace("$", "").replace("%", "").replace(" ", "").replace(".", "").replace(",", ".")
    try: return float(s)
    except: return 0.0

def dia_to_date(day_val, year_val):
    if pd.isna(day_val): return pd.NaT
    s_val = str(day_val).strip().lower()
    # Filtro de palabras que no son fechas
    if s_val in ["dia", "total", "none", "", "mes", "semana del año"]: return pd.NaT
    
    # Intento 1: Conversión directa de fecha
    dt = pd.to_datetime(day_val, dayfirst=True, errors="coerce")
    
    # Intento 2: Si es un número (ej. "1"), lo unimos con el año (asumiendo que pd.to_datetime lo maneja con el mes actual o lo fijamos)
    if pd.isna(dt):
        try:
            # Si el valor es un número entero (día del mes)
            day_num = int(float(day_val))
            if 1 <= day_num <= 31:
                # Intentamos construir una fecha mínima para que pd.to_datetime lo procese
                dt = pd.to_datetime(f"{day_num}/01/{int(year_val)}", dayfirst=True, errors='coerce')
        except: pass

    # Si el año detectado es 1900, forzamos el año de los filtros
    if pd.notna(dt) and year_val and dt.year == 1900:
        try: return dt.replace(year=int(year_val))
        except: return dt
    return dt

# --- INTERFAZ STREAMLIT ---
st.title("📊 Transformación de Indicadores WM")
st.markdown("Sube tu archivo ZIP para procesar y agrupar los indicadores correctamente.")

uploaded_file = st.file_uploader("Cargar archivo ZIP con reportes Excel", type="zip")

if uploaded_file is not None:
    if st.button("🚀 Iniciar Transformación"):
        try:
            by_key = {}
            PERCENT_COLS = ["Armado a Tiempo","OTEA","NPS","NSG","Completitud","Same Day","N2H","Contactos Perfectos","Participación","Variación %","Reclamos","Desviación","TEP"]
            FINAL_SCHEMA = ["Local ID","Año","Mes","Semana","Dia","Pedidos Facturados"] + PERCENT_COLS
            
            with zipfile.ZipFile(uploaded_file, 'r') as z:
                all_files = [fn for fn in z.namelist() if fn.lower().endswith(".xlsx") and "indicadores" in fn.lower()]
                
                if not all_files:
                    st.error("No se encontraron archivos válidos dentro del ZIP.")
                else:
                    progress_bar = st.progress(0)
                    for i, fn in enumerate(all_files):
                        with z.open(fn) as f:
                            content = f.read()
                            # 1. Metadatos (ID Local, Semana, Año)
                            meta_wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
                            ws = meta_wb["Export"] if "Export" in meta_wb.sheetnames else meta_wb.worksheets[0]
                            meta = {}
                            for r in range(1, 50):
                                v = ws.cell(r, 1).value
                                if isinstance(v, str) and "filtros aplicados:" in v.lower():
                                    meta = parse_filters_text(v)
                                    break
                            meta_wb.close()
                            
                            # 2. DETECCIÓN DINÁMICA DE TABLA (Solución al error 'Series')
                            df_raw = pd.read_excel(io.BytesIO(content), header=None)
                            header_idx = 0
                            for idx, row in df_raw.iterrows():
                                # Unimos la fila en un solo texto para buscar las palabras clave
                                row_text = " ".join(row.dropna().astype(str)).lower()
                                if "kpi" in row_text or "año" in row_text:
                                    header_idx = idx
                                    break
                            
                            # 3. Lectura real desde la fila detectada
                            df = pd.read_excel(io.BytesIO(content), skiprows=header_idx)
                            if df.empty: continue

                            # Forzamos nombres de columnas por posición (Blindaje)
                            df.columns.values[0] = "Año_Orig"
                            df.columns.values[1] = "Mes"
                            df.columns.values[2] = "Semana"
                            df.columns.values[3] = "Dia"
                            
                            # Limpieza de basura (Filas de Total o encabezados repetidos)
                            df = df[df["Dia"].notna()].copy()
                            df = df[~df["Dia"].astype(str).str.lower().str.contains("total|dia|mes|semana", na=False)]
                            
                            # Asignar Local ID y Año desde metadatos
                            anio_meta = meta.get("Año", 2026)
                            local_id = meta.get("Local ID", "Desconocido")
                            df.insert(0, "Local ID", local_id)
                            df["Año"] = anio_meta
                            
                            # PROCESAR FECHAS (DIA)
                            df["Dia"] = df.apply(lambda row: dia_to_date(row["Dia"], row["Año"]), axis=1)
                            
                            # Eliminar filas donde la fecha falló (limpieza final)
                            df = df[df["Dia"].notna()].copy()

                            # Convertir indicadores y limpiar formatos
                            for c in df.columns:
                                if c not in ["Local ID", "Año", "Mes", "Semana", "Dia"]:
                                    df[c] = df[c].apply(to_number)
                                    if c in PERCENT_COLS and df[c].gt(1).any():
                                        df[c] = df[c] / 100.0
                            
                            # Asegurar el esquema final
                            for c in FINAL_SCHEMA:
                                if c not in df.columns: df[c] = 0.0
                            
                            df_final = df[FINAL_SCHEMA].copy()
                            
                            # Clave para agrupar: Local + Año + Semana
                            key = (str(local_id), str(int(anio_meta)), str(meta.get("Semana", "XX")))
                            
                            if key not in by_key:
                                by_key[key] = df_final
                            else:
                                by_key[key] = pd.concat([by_key[key], df_final]).drop_duplicates(subset=["Dia", "Local ID"])
                        
                        progress_bar.progress((i + 1) / len(all_files))

            # --- GENERACIÓN DE SALIDA ---
            if by_key:
                out_zip = io.BytesIO()
                with zipfile.ZipFile(out_zip, "w") as new_z:
                    for (loc, yr, wk), fdf in by_key.items():
                        excel_buf = io.BytesIO()
                        with pd.ExcelWriter(excel_buf, engine='openpyxl', datetime_format="DD/MM/YYYY") as writer:
                            fdf.to_excel(writer, index=False)
                        new_z.writestr(f"Local_{loc}_Año{yr}_S{wk}.xlsx", excel_buf.getvalue())
                
                st.success(f"✅ Procesados {len(by_key)} locales con éxito.")
                st.download_button("📥 Descargar Resultados", out_zip.getvalue(), "Reportes_WM.zip", "application/zip")
                st.dataframe(df_final.head(10)) 
            else:
                st.warning("No se encontraron datos válidos.")
        except Exception as e:
            st.error(f"❌ Error crítico: {e}")
